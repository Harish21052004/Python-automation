import openpyxl 

workbook = openpyxl.load_workbook("./xlsheeet/dataentry.xlsx")
sheet = workbook['details']
total_rows = sheet.max_row
total_col = sheet.max_column

print(total_rows)
print(total_col)

for i in range(1,total_rows+1):
    for j in range(1, total_col+1):
        print(sheet.cell(i,j).value, end="  ")
    print()       